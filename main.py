import openpyxl
import re

class Client:

    def __init__(self):
        self.start = 3
        self.stop = 7
        self.sex = False  # female
        self.declination = True
        self.square = 1
        self.email = ''
        self.addr = ''
        self.fio_list = []
        self.tsg = ''

        self.wb = openpyxl.load_workbook('E:/kp/fio.xlsx')
        self.sheet = self.wb['fio']

        self.kpa = openpyxl.load_workbook('E:/kp/kpa.xlsx')
        self.sheet_A = self.kpa['kp']

        self.kpm = openpyxl.load_workbook('E:/kp/kpm.xlsx')
        self.sheet_M = self.kpm['kp']


    def parse_file(self, position):
        sex_regex = re.compile(r'вич$')

        self.fio_list = self.sheet['A'+str(position)].value.split(' ') #FIO
        self.addr = self.sheet['B'+str(position)].value
        self.email = self.sheet['D' + str(position)].value
        self.square = self.sheet['F' + str(position)].value
        self.tsg = self.sheet['H' + str(position)].value
        if len(self.fio_list) > 2: # определение пола
            self.sex = bool(sex_regex.search(self.fio_list[1]))
        if len(self.fio_list) > 3:
            self.surname_declination(self.fio_list[2], self.sex)

    def surname_declination(self, surname, sex):
        surname_regex = re.compile(r'(ых$)|(их$)|(е$)|(и$)|(о$)|(у$)|(ы$)|(э$)|(ю$)')
        not_declination = bool(surname_regex.search(surname))
        consonants = ['б', 'в', 'г', 'д', 'ж', 'й', 'з', 'к', 'л', 'м', 'н', 'п', 'р', 'с', 'т', 'ф', 'х', 'ц', 'ч', 'ш', 'щ']
        self.declination = True

        if not_declination == True:
            self.declination = False
        else:
            if sex == False: #женщина
                if consonants.count(surname[-1]) > 0:
                    self.declination = False
                else:
                    self.declination = True

    def recipient(self):
        if self.declination == False and len(self.fio_list) > 2:
            return self.fio_list[2]
        if self.sex == False:
            if self.declination == False and len(self.fio_list) > 2:
                return self.fio_list[2]
            else:
                surname_decl_re1 = re.compile(r'ая$')
                surname_decl_re2 = re.compile(r'ва$')
                if bool(surname_decl_re1.search(self.fio_list[2])):
                    return self.fio_list[2][:-2] + 'ой'
                elif bool(surname_decl_re2.search(self.fio_list[2])):
                    return self.fio_list[2][:-1] + 'ой'
                else:
                    return self.fio_list[2]
        else:
            surname_decl_re = re.compile(r'й$')
            if bool(surname_decl_re.search(self.fio_list[2])):
                return self.fio_list[2][:-2] + 'ому'
            else:
                return self.fio_list[2] + 'у'


    def make_kp_andrey(self):
        self.sheet_A['F7'].value = self.tsg
        self.sheet_A['F9'].value = self.recipient() + ' ' + self.fio_list[0][0] + '. ' + self.fio_list[1][0] + '.'
        self.sheet_A['F10'].value = 'email: ' + self.email
        if not self.sex:
            self.sheet_A['B18'].value = 'Уважаемая'
        else:
            self.sheet_A['B18'].value = 'Уважаемый'
        if len(self.fio_list) > 2:
            self.sheet_A['E18'].value = self.fio_list[0] + ' ' + self.fio_list[1] + '!'
        elif len(self.fio_list) > 1:
            self.sheet_A['E18'].value = self.fio_list[0]+'!'
        else:
            self.sheet_A['E18'].value = ''
        self.sheet_A['E21'].value = self.addr
        self.sheet_A['A22'].value = 'Площадь жилых помещений ~'+str(self.square)+' м. кв.'

        lk = float(self.square) * 1.9
        pt = float(self.square) * 1.8
        mp = float(self.square) * 1.4

        if float(self.square) < 15000:
            lk = float(self.square) * 2.05
            pt = float(self.square) * 1.85
            mp = float(self.square) * 1.45

        if float(self.square) < 12000:
            lk = float(self.square) * 2.05
            pt = float(self.square) * 1.9
            mp = float(self.square) * 1.5

        if float(self.square) < 11000:
            lk = float(self.square) * 2.07
            pt = float(self.square) * 1.92
            mp = float(self.square) * 1.6

        self.sheet_A['G52'].value = lk
        self.sheet_A['E53'].value = pt
        self.sheet_A['E54'].value = mp

        self.kpa.save('E:/kp/AN_' + self.tsg.replace('"', '') + '.xlsx')
    def make_kp_maxim(self):
        self.sheet_M['F7'].value = self.tsg
        self.sheet_M['F9'].value = self.recipient() + ' ' + self.fio_list[0][0] + '. ' + self.fio_list[1][0] + '.'
        self.sheet_M['F10'].value = 'email: ' + self.email
        if not self.sex:
            self.sheet_M['B18'].value = 'Уважаемая'
        else:
            self.sheet_M['B18'].value = 'Уважаемый'
        if len(self.fio_list) > 2:
            self.sheet_M['E18'].value = self.fio_list[0] + ' ' + self.fio_list[1] + '!'
        elif len(self.fio_list) > 1:
            self.sheet_M['E18'].value = self.fio_list[0] + '!'
        else:
            self.sheet_M['E18'].value = ''
        self.sheet_M['E21'].value = self.addr
        self.sheet_M['A22'].value = 'Площадь жилых помещений ~' + str(self.square) + ' м. кв.'

        lk = float(self.square) * 1.9
        pt = float(self.square) * 1.8
        mp = float(self.square) * 1.4

        if float(self.square) < 15000:
            lk = float(self.square) * 2.05
            pt = float(self.square) * 1.85
            mp = float(self.square) * 1.45

        if float(self.square) < 12000:
            lk = float(self.square) * 2.05
            pt = float(self.square) * 1.9
            mp = float(self.square) * 1.5

        if float(self.square) < 11000:
            lk = float(self.square) * 2.07
            pt = float(self.square) * 1.92
            mp = float(self.square) * 1.6

        self.sheet_M['G52'].value = lk
        self.sheet_M['E53'].value = pt
        self.sheet_M['E54'].value = mp

        self.kpm.save('E:/kp/maxim/' + self.tsg.replace('"', '') + '.xlsx')

if __name__ == '__main__':
    parser = Client()
    andr = True
    for position in range(183, 186):
        parser.parse_file(position=position)
        print(parser.recipient(), parser.fio_list)
        if andr:
            parser.make_kp_andrey()
            andr = not andr
        else:
            parser.make_kp_maxim()
            andr = not andr


#"Коммерческое предложение для"
#"Коммерческое предложение для"
#"Коммерческое предложение Уборка МКД"

По устному запросу направляю на рассмотрение коммерческое предложение по уборке МКД.
На возникшие вопросы готов ответить по телефону или по почте.

Продолжая телефонный разговор, высылаю Вам на рассмотрение коммерческое предложение по
Уборке мест общего пользования, придомовой территории МДК, обслуживанию мусоропроводов.

Направляю Вам на рассмотрение коммерческое предложение по уборке мест общего пользования
и придомовой территории МКД. Файл во вложении.

